#!/usr/bin/env python3
"""
HARVEST CONVERSATIONS
=====================
Scrape ALL Claude.ai conversation URLs and build Excel catalogs.

This script uses browser automation to:
1. Log into Claude.ai
2. Scroll through the conversation sidebar
3. Extract ALL conversation URLs and titles
4. Save to Excel files in batches

USAGE:
  python harvest-conversations.py              # Interactive mode
  python harvest-conversations.py --headless   # Headless browser
  python harvest-conversations.py --batch 50   # 50 conversations per xlsx

REQUIREMENTS:
  pip install playwright openpyxl anthropic python-dotenv
  playwright install chromium
"""

import json
import asyncio
import argparse
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import os

# Load .env
env_paths = [
    Path(__file__).parent / ".env",
    Path(__file__).parent.parent.parent.parent / ".env",
    Path.home() / ".env",
]
for p in env_paths:
    if p.exists():
        load_dotenv(p)
        break

OUTPUT_DIR = Path(__file__).parent / "conversation-harvest"
OUTPUT_DIR.mkdir(exist_ok=True)


async def harvest_with_playwright(headless: bool = False, batch_size: int = 20):
    """Use Playwright to scrape conversation URLs from Claude.ai."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("ERROR: pip install playwright && playwright install chromium")
        return []

    print("="*60)
    print("HARVESTING CONVERSATIONS FROM CLAUDE.AI")
    print("="*60)
    print(f"Mode: {'headless' if headless else 'visible browser'}")
    print(f"Batch size: {batch_size} conversations per xlsx")
    print()

    conversations = []

    async with async_playwright() as p:
        # Launch browser
        browser = await p.chromium.launch(headless=headless)
        context = await browser.new_context()

        # Check for saved session
        session_file = OUTPUT_DIR / "claude-session.json"
        if session_file.exists():
            print("Loading saved session...")
            with open(session_file) as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)

        page = await context.new_page()

        # Go to Claude.ai
        print("\nNavigating to claude.ai...")
        await page.goto("https://claude.ai")
        await page.wait_for_timeout(3000)

        # Check if logged in
        current_url = page.url
        if "login" in current_url or "auth" in current_url:
            print("\n" + "="*60)
            print("LOGIN REQUIRED")
            print("="*60)
            print("Please log in manually in the browser window.")
            print("The script will continue once you're logged in.")
            print()

            # Wait for login (check for conversation sidebar)
            try:
                await page.wait_for_selector('[data-testid="conversation-list"]', timeout=300000)
            except:
                # Alternative selector
                await page.wait_for_url("**/chat/**", timeout=300000)

            print("Login detected! Saving session...")
            cookies = await context.cookies()
            with open(session_file, "w") as f:
                json.dump(cookies, f)

        # Now we're logged in, scroll and collect conversations
        print("\nCollecting conversations...")

        # Wait for sidebar to load
        await page.wait_for_timeout(2000)

        # Scroll to load all conversations
        scroll_count = 0
        prev_count = 0

        while True:
            # Get all conversation links
            links = await page.query_selector_all('a[href^="/chat/"]')

            current_count = len(links)
            if current_count == prev_count and scroll_count > 0:
                # No new conversations loaded, we're done
                break

            prev_count = current_count
            scroll_count += 1

            print(f"  Scroll {scroll_count}: Found {current_count} conversations...")

            # Scroll down in sidebar
            sidebar = await page.query_selector('[data-testid="conversation-list"]')
            if sidebar:
                await sidebar.evaluate("el => el.scrollTop = el.scrollHeight")
            else:
                # Try alternative scrolling
                await page.keyboard.press("End")

            await page.wait_for_timeout(1000)

            # Safety limit
            if scroll_count > 100:
                print("  Reached scroll limit (100)")
                break

        # Extract all conversation data
        print(f"\nExtracting {len(links)} conversations...")

        for i, link in enumerate(links):
            try:
                href = await link.get_attribute("href")
                title_el = await link.query_selector("span, div")
                title = await title_el.inner_text() if title_el else "Untitled"

                # Clean up URL
                if href.startswith("/"):
                    url = f"https://claude.ai{href}"
                else:
                    url = href

                # Extract conversation ID
                conv_id = href.split("/")[-1] if href else None

                conversations.append({
                    "url": url,
                    "title": title.strip()[:100],  # Limit title length
                    "id": conv_id,
                    "index": i + 1,
                })

                if (i + 1) % 50 == 0:
                    print(f"  Processed {i + 1}/{len(links)}")

            except Exception as e:
                print(f"  Error on conversation {i}: {e}")

        await browser.close()

    print(f"\n✓ Collected {len(conversations)} conversations")
    return conversations


def save_to_excel(conversations: list[dict], batch_size: int = 20):
    """Save conversations to Excel files in batches."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: pip install openpyxl")
        return

    if not conversations:
        print("No conversations to save")
        return

    # Sort by index (oldest first, or however they were collected)
    conversations = sorted(conversations, key=lambda x: x.get("index", 0))

    # Create batches
    batches = []
    for i in range(0, len(conversations), batch_size):
        batches.append(conversations[i:i + batch_size])

    print(f"\nSaving {len(conversations)} conversations to {len(batches)} Excel files...")

    for batch_num, batch in enumerate(batches, 1):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Batch {batch_num}"

        # Headers
        headers = ["#", "URL", "Title", "ID"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Widen URL and Title columns
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 50

        # Data rows
        for row_num, conv in enumerate(batch, 2):
            ws.cell(row=row_num, column=1, value=conv.get("index", row_num - 1))

            # Add URL as hyperlink
            url = conv.get("url", "")
            title = conv.get("title", "Untitled")
            cell = ws.cell(row=row_num, column=2, value=url)
            cell.hyperlink = url
            cell.style = "Hyperlink"

            ws.cell(row=row_num, column=3, value=title)
            ws.cell(row=row_num, column=4, value=conv.get("id", ""))

        # Save
        filename = f"claude-conversations-batch{batch_num:02d}.xlsx"
        filepath = OUTPUT_DIR / filename
        wb.save(filepath)
        print(f"  ✓ {filename}: {len(batch)} conversations")

    # Also save master JSON
    master_path = OUTPUT_DIR / "all-conversations.json"
    with open(master_path, "w") as f:
        json.dump(conversations, f, indent=2)
    print(f"\n✓ Master JSON: {master_path}")

    # Summary
    print(f"\n" + "="*60)
    print("HARVEST COMPLETE")
    print("="*60)
    print(f"Total conversations: {len(conversations)}")
    print(f"Excel files: {len(batches)}")
    print(f"Output folder: {OUTPUT_DIR}")


async def harvest_via_api():
    """
    Alternative: Use Claude API to help categorize conversations.
    This requires conversation exports to already exist.
    """
    import anthropic

    client = anthropic.Anthropic()

    print("="*60)
    print("API-ASSISTED CATEGORIZATION")
    print("="*60)

    # Look for existing conversation JSON exports
    exports_dir = OUTPUT_DIR / "exports"
    if not exports_dir.exists():
        print(f"\nNo exports folder found at {exports_dir}")
        print("Export your conversations from Claude.ai first.")
        return

    conversations = []

    for export_file in sorted(exports_dir.glob("*.json")):
        print(f"Processing {export_file.name}...")

        with open(export_file) as f:
            conv = json.load(f)

        # Extract first message to determine topic
        messages = conv.get("messages", [])
        first_user_msg = ""
        for msg in messages[:3]:
            if msg.get("role") == "user":
                first_user_msg = msg.get("content", "")[:500]
                break

        if not first_user_msg:
            continue

        # Use Claude to categorize
        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=200,
                messages=[{
                    "role": "user",
                    "content": f"""Given this conversation start, provide a short title (max 10 words) and topic tags (comma-separated).

CONVERSATION START:
{first_user_msg}

Output JSON only:
{{"title": "...", "tags": "tag1, tag2, tag3"}}"""
                }]
            )

            result_text = response.content[0].text
            # Parse JSON
            import re
            match = re.search(r'\{.*\}', result_text, re.DOTALL)
            if match:
                result = json.loads(match.group())
                conversations.append({
                    "file": export_file.name,
                    "title": result.get("title", "Untitled"),
                    "tags": result.get("tags", ""),
                    "id": conv.get("id", export_file.stem),
                })
                print(f"  ✓ {result.get('title', 'Untitled')}")

        except Exception as e:
            print(f"  ✗ Error: {e}")

    # Save categorized list
    if conversations:
        out_path = OUTPUT_DIR / "categorized-conversations.json"
        with open(out_path, "w") as f:
            json.dump(conversations, f, indent=2)
        print(f"\n✓ Saved {len(conversations)} categorized conversations to {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Harvest Claude.ai conversations")
    parser.add_argument("--headless", action="store_true",
                        help="Run browser in headless mode")
    parser.add_argument("--batch", type=int, default=20,
                        help="Conversations per Excel file (default: 20)")
    parser.add_argument("--api", action="store_true",
                        help="Use API to categorize existing exports instead of scraping")
    args = parser.parse_args()

    if args.api:
        asyncio.run(harvest_via_api())
    else:
        conversations = asyncio.run(harvest_with_playwright(
            headless=args.headless,
            batch_size=args.batch
        ))
        if conversations:
            save_to_excel(conversations, batch_size=args.batch)


if __name__ == "__main__":
    main()
