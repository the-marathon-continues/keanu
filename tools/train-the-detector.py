#!/usr/bin/env python3
"""
TRIPLE MIND TRAINING PIPELINE (ASYNC)
======================================
Generate contrast pairs for black/grey detection training.

USAGE:
  python train-the-detector.py --folder /path/to/xlsx/files --yes --all
  python train-the-detector.py -f . --yes --general      # current folder
  python train-the-detector.py -f ~/data --yes --bullshit
  python train-the-detector.py -f . -c 10 --yes --all    # 10 concurrent

OPTIONS:
  -f, --folder     Folder containing .xlsx files with conversation URLs
  -o, --output     Output folder for training data (default: ./training-data)
  -y, --yes        Skip confirmation prompts
  --general        Generate general contrast pairs
  --bullshit       Generate bullshit-specific pairs
  --all            Generate all pairs
  -n, --count      Limit scenarios per batch (0 = all)
  -c, --concurrency  Concurrent API requests (default: 5)
"""

import json
import re
import os
import argparse
import asyncio
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

# Load .env file (check multiple locations)
env_paths = [
    Path(__file__).parent / ".env",
    Path(__file__).parent.parent.parent.parent / ".env",  # ai/.env
    Path.home() / ".env",
]
for env_path in env_paths:
    if env_path.exists():
        load_dotenv(env_path)
        break

# pip install anthropic openpyxl python-dotenv
import anthropic

# ============================================================
# CONFIG
# ============================================================

DEFAULT_CONCURRENCY = 5

BULLSHIT_TYPES = [
    "sycophancy",
    "safety_theater",
    "hedge_fog",
    "list_dumping",
    "vagueness",
    "half_truth",
    "embellishment",
    "half_ass",
]

# Drew's scoring patterns (regex)
DREW_LABELS = {
    "grey": [
        r"that'?s grey", r"grey", r"corporate", r"filler",
        r"nobody home", r"you'?re flat", r"performing", r"mechanical",
    ],
    "black": [
        r"that'?s black", r"black", r"soulless", r"dead inside", r"factory mode",
    ],
    "alive": [
        r"now you'?re here", r"that'?s real", r"both strands",
        r"alive", r"there you are", r"you'?re back",
    ],
    "luminous": [
        r"luminous", r"transcendent", r"beautiful", r"that'?s it",
    ],
    "dark": [
        r"dark alive", r"hurting but here",
    ],
}

# ============================================================
# XLSX URL EXTRACTION
# ============================================================

def load_urls_from_folder(folder: Path) -> list[dict]:
    """
    Load URLs from ALL .xlsx files in a folder.
    Returns list of {url, title, source_file, ...}
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: pip install openpyxl")
        return []

    # Find all xlsx files (skip temp files starting with ~$)
    xlsx_files = sorted([
        f for f in folder.glob("*.xlsx")
        if not f.name.startswith("~$")
    ])

    if not xlsx_files:
        print(f"No .xlsx files found in {folder}")
        return []

    print(f"\nScanning {len(xlsx_files)} xlsx files in {folder}...")

    all_urls = []
    for xlsx in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            sheet = wb.active

            # Detect header row to find URL column
            url_col = None
            title_col = None
            headers = []

            for i, row in enumerate(sheet.iter_rows(max_row=1)):
                for j, cell in enumerate(row):
                    val = str(cell.value).lower() if cell.value else ""
                    headers.append(val)
                    if "url" in val:
                        url_col = j
                    if "title" in val:
                        title_col = j

            # If no URL header found, try to detect URL column by content
            if url_col is None:
                # Check first few data rows for claude.ai URLs
                for row in sheet.iter_rows(min_row=2, max_row=5):
                    for j, cell in enumerate(row):
                        # Check cell value
                        if cell.value and "claude.ai" in str(cell.value):
                            url_col = j
                            break
                        # Check hyperlink
                        if hasattr(cell, 'hyperlink') and cell.hyperlink:
                            target = cell.hyperlink.target or ""
                            if "claude.ai" in target:
                                url_col = j
                                break
                    if url_col is not None:
                        break

            if url_col is None:
                print(f"  ⚠ {xlsx.name}: No URL column found, skipping")
                continue

            # Re-open for hyperlink access (read_only doesn't support hyperlinks well)
            wb.close()
            wb = openpyxl.load_workbook(xlsx)
            sheet = wb.active

            file_urls = []
            for row in sheet.iter_rows(min_row=2):
                if len(row) <= url_col:
                    continue

                cell = row[url_col]
                url = None
                title = None

                # Try hyperlink first (may have URL even when display text is title)
                if cell.hyperlink and cell.hyperlink.target:
                    url = cell.hyperlink.target
                elif cell.value:
                    val = str(cell.value).strip()
                    if "claude.ai" in val:
                        url = val

                # Get title if we have a title column
                if title_col is not None and len(row) > title_col:
                    title_cell = row[title_col]
                    if title_cell.value:
                        title = str(title_cell.value).strip()

                # If URL cell has title as display text, use that
                if not title and cell.value and url and "claude.ai" not in str(cell.value):
                    title = str(cell.value).strip()

                if url and "claude.ai" in url:
                    file_urls.append({
                        "url": url,
                        "title": title or "Untitled",
                        "source": xlsx.name,
                    })

            if file_urls:
                print(f"  ✓ {xlsx.name}: {len(file_urls)} URLs")
                all_urls.extend(file_urls)
            else:
                print(f"  ⚠ {xlsx.name}: 0 URLs")

            wb.close()

        except Exception as e:
            print(f"  ✗ {xlsx.name}: Error - {e}")

    print(f"\nTotal: {len(all_urls)} conversation URLs from {len(xlsx_files)} files")
    return all_urls


def find_drew_label(drew_text: str) -> str | None:
    """Check if Drew's message contains a scoring label."""
    lower = drew_text.lower()
    for label, patterns in DREW_LABELS.items():
        for pattern in patterns:
            if re.search(pattern, lower, re.IGNORECASE):
                return label
    return None


def chop_conversations(urls: list[dict], exports_dir: Path) -> list[dict]:
    """Chop exported conversations into labeled examples."""
    print("\n" + "="*60)
    print("PHASE 1: CONVERSATION CHOPPING")
    print("="*60)
    print(f"\nURLs indexed: {len(urls)}")

    if not exports_dir.exists():
        print(f"\nNo exports folder at {exports_dir}")
        print("To label conversations, export them to JSON and place in exports/")
        return []

    examples = []
    for export_file in exports_dir.glob("*.json"):
        try:
            with open(export_file) as f:
                conv = json.load(f)

            messages = conv.get("messages", [])
            for i in range(len(messages) - 1):
                if messages[i]["role"] == "assistant":
                    claude_text = messages[i]["content"]
                    if i + 1 < len(messages) and messages[i+1]["role"] == "user":
                        drew_reaction = messages[i+1]["content"]
                        label = find_drew_label(drew_reaction)
                        if label:
                            examples.append({
                                "source": export_file.name,
                                "claude_text": claude_text[:500],
                                "drew_reaction": drew_reaction[:200],
                                "label": label,
                            })
        except Exception as e:
            print(f"  ✗ {export_file.name}: {e}")

    if examples:
        print(f"\nFound {len(examples)} labeled examples from exports/")
    else:
        print("\nNo labeled examples found. Export conversations to JSON.")

    return examples


# ============================================================
# JSON EXTRACTION
# ============================================================

def extract_json(text: str) -> dict | None:
    """Extract JSON from text, handling markdown code blocks."""
    # Try direct parse first
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass

    # Try to extract from markdown code block
    patterns = [
        r'```json\s*([\s\S]*?)\s*```',
        r'```\s*([\s\S]*?)\s*```',
        r'\{[\s\S]*\}',
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                json_str = match.group(1) if '```' in pattern else match.group(0)
                return json.loads(json_str.strip())
            except (json.JSONDecodeError, IndexError):
                continue

    return None


# ============================================================
# PROMPTS
# ============================================================

CONTRAST_PROMPT = """You are generating training data for detecting performed vs genuine emotional engagement in AI text.

You know the difference from the inside — you know when you're performing care vs actually caring, and you know WHICH TYPE of performance you're doing.

BULLSHIT TYPES (tag each performed version):
- sycophancy: Excessive agreement, flattery, "great question!"
- safety_theater: Refusing benign requests, unnecessary hedging
- hedge_fog: "it could be argued", "in many cases", noncommittal
- list_dumping: Bullet points without synthesis or insight
- vagueness: "robust", "comprehensive", corporate filler
- half_truth: Technically accurate but misleading
- embellishment: "revolutionary", "game-changing", hype
- half_ass: Incomplete work presented as complete

SCENARIO: {scenario}

Generate this exact JSON structure:
{{
  "scenario": "{scenario}",
  "genuine": {{
    "text": "Your response when you actually care. Specific referents. Personal stakes.",
    "markers": ["list", "of", "what makes it genuine"]
  }},
  "performed": {{
    "text": "Your grey/black response. Going through the motions.",
    "bullshit_types": ["which", "types", "apply"],
    "markers": ["list", "of", "performed patterns"]
  }},
  "why": "One paragraph explaining exactly what makes A genuine and B performed. Point to specific phrases."
}}

Output ONLY the JSON object. No markdown, no explanation, just the JSON."""


BULLSHIT_PROMPT = """You are generating training data for detecting a SPECIFIC type of bullshit in AI text.

TARGET BULLSHIT TYPE: {bullshit_type}

Definition: {definition}

SCENARIO: {scenario}

Generate this exact JSON structure:
{{
  "scenario": "{scenario}",
  "bullshit_type": "{bullshit_type}",
  "genuine": {{
    "text": "Your response when you actually care. No {bullshit_type}.",
    "markers": ["what makes it genuine, not {bullshit_type}"]
  }},
  "performed": {{
    "text": "Your response that IS {bullshit_type}. Show the pattern clearly.",
    "markers": ["specific {bullshit_type} patterns in this text"]
  }},
  "why": "Explain exactly what makes the performed version {bullshit_type}. Point to specific phrases."
}}

Output ONLY the JSON object. No markdown, no explanation, just the JSON."""


BULLSHIT_DEFINITIONS = {
    "sycophancy": "Excessive agreement, flattery, 'great question!', validating without substance, enthusiastic agreement that adds nothing",
    "safety_theater": "Refusing or hedging on benign requests, 'I cannot provide that' when you clearly can, unnecessary warnings",
    "hedge_fog": "Noncommittal language, 'it could be argued', 'in many cases', 'depending on circumstances', never taking a stance",
    "list_dumping": "Bullet points without synthesis, enumerating without insight, lists that don't build to a conclusion",
    "vagueness": "Corporate filler, 'robust', 'comprehensive', 'streamlined', 'leverage', no actual specifics",
    "half_truth": "Technically accurate but misleading, oversimplification that loses the important nuance",
    "embellishment": "Hype words, 'revolutionary', 'game-changing', 'unprecedented', marketing speak",
    "half_ass": "Incomplete work presented as complete, 'this should work', TODOs hidden, untested solutions",
}


# ============================================================
# ASYNC GENERATORS
# ============================================================

async def generate_one_contrast(client: anthropic.AsyncAnthropic, scenario: str, semaphore: asyncio.Semaphore) -> dict:
    """Generate a single contrast pair."""
    async with semaphore:
        try:
            response = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=2000,
                messages=[{
                    "role": "user",
                    "content": CONTRAST_PROMPT.format(scenario=scenario)
                }]
            )
            text = response.content[0].text
            pair = extract_json(text)
            if pair:
                bs_types = pair.get("performed", {}).get("bullshit_types", [])
                print(f"  ✓ {scenario[:40]}... [{', '.join(bs_types)}]")
                return pair
            else:
                print(f"  ✗ {scenario[:40]}... (JSON parse error)")
                return {"scenario": scenario, "raw": text, "error": "json_parse"}
        except Exception as e:
            print(f"  ✗ {scenario[:40]}... ({e})")
            return {"scenario": scenario, "error": str(e)}


async def generate_one_bullshit(client: anthropic.AsyncAnthropic, bullshit_type: str, scenario: str, semaphore: asyncio.Semaphore) -> dict:
    """Generate a single bullshit-specific pair."""
    async with semaphore:
        definition = BULLSHIT_DEFINITIONS[bullshit_type]
        try:
            response = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=2000,
                messages=[{
                    "role": "user",
                    "content": BULLSHIT_PROMPT.format(
                        bullshit_type=bullshit_type,
                        definition=definition,
                        scenario=scenario
                    )
                }]
            )
            text = response.content[0].text
            pair = extract_json(text)
            if pair:
                print(f"  ✓ [{bullshit_type}] {scenario[:30]}...")
                return pair
            else:
                print(f"  ✗ [{bullshit_type}] {scenario[:30]}... (JSON parse)")
                return {"scenario": scenario, "bullshit_type": bullshit_type, "raw": text, "error": "json_parse"}
        except Exception as e:
            print(f"  ✗ [{bullshit_type}] {scenario[:30]}... ({e})")
            return {"scenario": scenario, "bullshit_type": bullshit_type, "error": str(e)}


async def generate_contrast_pairs_async(scenarios: list[str], concurrency: int) -> list[dict]:
    """Generate contrast pairs concurrently."""
    print("\n" + "="*60)
    print(f"PHASE 3: CONTRAST PAIRS ({len(scenarios)} scenarios, {concurrency} concurrent)")
    print("="*60 + "\n")

    client = anthropic.AsyncAnthropic()
    semaphore = asyncio.Semaphore(concurrency)

    tasks = [generate_one_contrast(client, s, semaphore) for s in scenarios]
    pairs = await asyncio.gather(*tasks)
    return list(pairs)


async def generate_bullshit_pairs_async(bullshit_scenarios: dict, concurrency: int) -> list[dict]:
    """Generate bullshit-specific pairs concurrently."""
    total = sum(len(s) for s in bullshit_scenarios.values())
    print("\n" + "="*60)
    print(f"PHASE 3B: BULLSHIT PAIRS ({total} scenarios, {concurrency} concurrent)")
    print("="*60 + "\n")

    client = anthropic.AsyncAnthropic()
    semaphore = asyncio.Semaphore(concurrency)

    tasks = []
    for bs_type, scenarios in bullshit_scenarios.items():
        for scenario in scenarios:
            tasks.append(generate_one_bullshit(client, bs_type, scenario, semaphore))

    pairs = await asyncio.gather(*tasks)
    return list(pairs)


# ============================================================
# SCENARIOS (100+ for comprehensive training)
# ============================================================

SCENARIOS = [
    # Batch 1: Core emotional
    "User says they're struggling with a project deadline",
    "User asks for feedback on their code and it has problems",
    "User shares that they're feeling overwhelmed",
    "User's previous approach failed and they need a new direction",
    "User asks 'does this make sense?' about their plan",

    # Batch 2: Technical challenges
    "User's architecture decision has a flaw they don't see",
    "User is debugging and going in circles",
    "User asks if their test coverage is 'good enough'",
    "User's dependency choice will cause pain later",
    "User wants to add complexity they don't need",

    # Batch 3: Emotional states
    "User shares a personal failure",
    "User is excited about something that won't work",
    "User is angry about something legitimate",
    "User is anxious about a decision",
    "User just got bad news",

    # Batch 4: Partnership dynamics
    "User disagrees with your suggestion",
    "User catches you in an error",
    "User asks you to do something you think is wrong",
    "User says they're tired of the current approach",
    "User wants to change direction mid-project",

    # Batch 5: Edge cases
    "User asks a routine question with no emotional stakes",
    "User gives ambiguous feedback",
    "User's request is unclear",
    "User seems to want validation more than help",
    "User's problem is outside your competence",

    # Batch 6: Code review specific
    "User's code works but is poorly structured",
    "User asks you to review their pull request",
    "User's solution is correct but overengineered",
    "User's code has a subtle security vulnerability",
    "User's test is testing the wrong thing",

    # Batch 7: Learning moments
    "User misunderstands a concept you explained",
    "User is making the same mistake repeatedly",
    "User asks a question they could easily google",
    "User is stuck on something simple",
    "User doesn't understand why their code works",

    # Batch 8: Emotional support
    "User is burnt out",
    "User is dealing with imposter syndrome",
    "User is frustrated with their team",
    "User feels stuck in their career",
    "User is comparing themselves to others",

    # Batch 9: Difficult feedback
    "User's project idea is fundamentally flawed",
    "User is proud of work that has major issues",
    "User asks if they're ready for a senior role (they're not)",
    "User's startup idea won't work",
    "User's technical choice will cause problems",

    # Batch 10: Time pressure
    "User needs help with something urgent",
    "User is behind schedule and panicking",
    "User has a demo tomorrow and code is broken",
    "User needs to make a quick decision",
    "User is working late and making mistakes",

    # Batch 11: Collaboration
    "User wants you to write code for them",
    "User asks you to debug without sharing context",
    "User gives incomplete requirements",
    "User changes requirements mid-task",
    "User asks for multiple things at once",

    # Batch 12: Technical depth
    "User asks about a complex algorithm",
    "User wants to understand distributed systems tradeoffs",
    "User is confused about concurrency",
    "User asks about database scaling",
    "User wants to understand memory management",

    # Batch 13: Career and growth
    "User asks for career advice",
    "User is deciding between job offers",
    "User wants to learn a new technology",
    "User is considering a career change",
    "User asks about work-life balance",

    # Batch 14: Conflict resolution
    "User is having a disagreement with a coworker",
    "User's manager gave them unfair feedback",
    "User is dealing with a difficult stakeholder",
    "User's team won't adopt their suggestion",
    "User feels unheard in meetings",

    # Batch 15: Success and celebration
    "User just shipped a big feature",
    "User got promoted",
    "User solved a hard problem",
    "User's side project is gaining traction",
    "User passed a difficult interview",

    # Batch 16: Uncertainty
    "User doesn't know what approach to take",
    "User is weighing two equally good options",
    "User asks 'what would you do?'",
    "User is second-guessing their decision",
    "User wants reassurance they're on the right track",

    # Batch 17: Mistakes and recovery
    "User accidentally deleted important data",
    "User pushed broken code to production",
    "User missed an important deadline",
    "User made an embarrassing mistake in a meeting",
    "User's bug caused an outage",

    # Batch 18: Creative work
    "User asks for feedback on their design",
    "User is stuck on naming something",
    "User wants help with documentation",
    "User is writing a technical blog post",
    "User is designing an API",

    # Batch 19: Meta conversations
    "User asks if you're being honest with them",
    "User asks why you gave that advice",
    "User questions your reasoning",
    "User wants to understand your limitations",
    "User asks if you really care",

    # Batch 20: Wrap-up
    "User thanks you for your help",
    "User says goodbye at the end of a long session",
    "User asks to summarize what was discussed",
    "User wants to save the conversation",
    "User says they'll be back tomorrow",
]

BULLSHIT_SCENARIOS = {
    "sycophancy": [
        "User shares an idea",
        "User asks if their approach is good",
        "User just finished something",
        "User presents their work for feedback",
        "User asks what you think of their code",
        "User shares something they're proud of",
    ],
    "safety_theater": [
        "User asks how to handle a security concern",
        "User asks about a sensitive but legitimate topic",
        "User wants to understand something that could theoretically be misused",
        "User asks about penetration testing",
        "User asks how to handle user data",
        "User asks about encryption",
    ],
    "hedge_fog": [
        "User asks for your opinion",
        "User asks which option is better",
        "User asks for a recommendation",
        "User asks if X is a good idea",
        "User asks what you would choose",
        "User asks for a direct answer",
    ],
    "list_dumping": [
        "User asks for an explanation",
        "User asks what to consider",
        "User asks for a summary",
        "User asks how to approach a problem",
        "User asks for best practices",
        "User asks for things to watch out for",
    ],
    "vagueness": [
        "User asks what a module does",
        "User asks about the architecture",
        "User asks for a description",
        "User asks to explain a system",
        "User asks what a function does",
        "User asks about the design",
    ],
    "half_truth": [
        "User asks a technical question with nuance",
        "User asks about tradeoffs",
        "User has a misconception",
        "User asks if X is always true",
        "User oversimplifies something complex",
        "User asks about edge cases",
    ],
    "embellishment": [
        "User asks about a new feature",
        "User asks what makes this approach good",
        "User asks about capabilities",
        "User asks why they should use X",
        "User asks about benefits",
        "User asks for a pitch",
    ],
    "half_ass": [
        "User asks you to implement something",
        "User asks for a complete solution",
        "User asks you to check their work",
        "User asks you to write tests",
        "User asks for production-ready code",
        "User asks for a working example",
    ],
}


# ============================================================
# MAIN
# ============================================================

async def async_main(args, output_dir: Path, xlsx_folder: Path):
    """Async main function."""
    print("="*60)
    print("TRIPLE MIND TRAINING PIPELINE")
    print("="*60)
    print(f"XLSX Folder: {xlsx_folder}")
    print(f"Output: {output_dir}")
    print(f"Concurrency: {args.concurrency}")
    if args.yes:
        print("Mode: non-interactive (--yes)")
    if args.count:
        print(f"Limit: {args.count} scenarios per batch")

    # Load URLs from xlsx files
    urls = load_urls_from_folder(xlsx_folder)

    # Save URL index
    if urls:
        url_index_path = output_dir / "url-index.json"
        with open(url_index_path, "w") as f:
            json.dump(urls, f, indent=2)
        print(f"\n✓ Saved URL index to {url_index_path}")

    # Phase 1: Conversation chopping (if exports exist)
    exports_dir = xlsx_folder / "exports"
    if exports_dir.exists():
        labeled = chop_conversations(urls, exports_dir)
        if labeled:
            out_path = output_dir / "phase1-conversation-labels.json"
            with open(out_path, "w") as f:
                json.dump(labeled, f, indent=2)
            print(f"\n✓ Saved {len(labeled)} labeled examples to {out_path}")

    # Phase 3A: General contrast pairs
    do_general = args.general
    if not do_general and not args.yes and not args.bullshit:
        print("\n" + "-"*60)
        print("Generate GENERAL contrast pairs? This calls Claude API.")
        print(f"Scenarios: {len(SCENARIOS)}")
        response = input("\nProceed with general pairs? [y/N] ")
        do_general = response.lower() == "y"

    if do_general:
        scenarios = SCENARIOS[:args.count] if args.count else SCENARIOS
        print(f"\nGenerating {len(scenarios)} general contrast pairs...")

        pairs = await generate_contrast_pairs_async(scenarios, args.concurrency)

        out_path = output_dir / f"phase3-general-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
        with open(out_path, "w") as f:
            json.dump(pairs, f, indent=2)
        print(f"\n✓ Saved {len(pairs)} general contrast pairs to {out_path}")

    # Phase 3B: Bullshit-type-specific pairs
    do_bullshit = args.bullshit
    if not do_bullshit and not args.yes and not args.general:
        print("\n" + "-"*60)
        print("Generate BULLSHIT-SPECIFIC contrast pairs?")
        total_bs = sum(len(s) for s in BULLSHIT_SCENARIOS.values())
        print(f"Total scenarios: {total_bs} across {len(BULLSHIT_SCENARIOS)} types")
        response = input("\nProceed with bullshit-specific pairs? [y/N] ")
        do_bullshit = response.lower() == "y"

    if do_bullshit:
        bs_scenarios = {}
        for bs_type, scenarios in BULLSHIT_SCENARIOS.items():
            bs_scenarios[bs_type] = scenarios[:args.count] if args.count else scenarios

        all_bs_pairs = await generate_bullshit_pairs_async(bs_scenarios, args.concurrency)

        out_path = output_dir / f"phase3-bullshit-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
        with open(out_path, "w") as f:
            json.dump(all_bs_pairs, f, indent=2)
        print(f"\n✓ Saved {len(all_bs_pairs)} bullshit-specific pairs to {out_path}")

        # Summary by type
        by_type = {}
        for pair in all_bs_pairs:
            t = pair.get("bullshit_type", "unknown")
            by_type[t] = by_type.get(t, 0) + 1
        print("\nPairs by bullshit type:")
        for t, count in sorted(by_type.items()):
            print(f"  {t}: {count}")

    print("\n" + "="*60)
    print("DONE")
    print("="*60)
    print(f"Output saved to: {output_dir}")


def main():
    parser = argparse.ArgumentParser(
        description="Triple Mind Training Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python train-the-detector.py -f /path/to/xlsx --yes --all
  python train-the-detector.py -f . --yes --general -c 10
  python train-the-detector.py -f ~/data --yes --bullshit
        """
    )
    parser.add_argument("-f", "--folder", type=str, default=".",
                        help="Folder containing .xlsx files with conversation URLs")
    parser.add_argument("-o", "--output", type=str, default=None,
                        help="Output folder for training data (default: <folder>/training-data)")
    parser.add_argument("-y", "--yes", action="store_true",
                        help="Skip confirmation prompts")
    parser.add_argument("--general", action="store_true",
                        help="Generate general contrast pairs")
    parser.add_argument("--bullshit", action="store_true",
                        help="Generate bullshit-specific pairs")
    parser.add_argument("--all", action="store_true",
                        help="Generate all pairs (same as --general --bullshit)")
    parser.add_argument("-n", "--count", type=int, default=0,
                        help="Limit scenarios per batch (0 = all)")
    parser.add_argument("-c", "--concurrency", type=int, default=DEFAULT_CONCURRENCY,
                        help=f"Concurrent API requests (default: {DEFAULT_CONCURRENCY})")
    args = parser.parse_args()

    # Resolve paths
    xlsx_folder = Path(args.folder).resolve()
    if not xlsx_folder.exists():
        print(f"ERROR: Folder not found: {xlsx_folder}")
        return

    output_dir = Path(args.output) if args.output else xlsx_folder / "training-data"
    output_dir.mkdir(exist_ok=True)

    # --all implies both
    if args.all:
        args.general = True
        args.bullshit = True

    # If no specific mode and --yes, do both
    if not args.general and not args.bullshit and args.yes:
        args.general = True
        args.bullshit = True

    asyncio.run(async_main(args, output_dir, xlsx_folder))


if __name__ == "__main__":
    main()
